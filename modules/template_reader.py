import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

def extract_tables(ws):
    dfs_tmp = {}
    for name, table in ws.tables.items():
        table_range = table.ref if hasattr(table, 'ref') else table
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        table = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        header = next(table)
        df = pd.DataFrame(table, columns=header)
        print(f"Extracted table {name} with {len(df)} rows")
        dfs_tmp[name] = df
    return dfs_tmp

def load_template(output_file_path):
    wb = load_workbook(output_file_path)
    mapping_ws = wb["Mapping"]
    item_ws = wb["Item"]
    return wb, mapping_ws, item_ws

def load_mechanical_template(output_file_path):
    wb = load_workbook(output_file_path)
    project_ws = wb["Project"]
    item_ws = wb["Item"]
    return wb, project_ws, item_ws