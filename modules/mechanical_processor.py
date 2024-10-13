import pandas as pd
import openpyxl as pyxl
from modules import mechanical_ref as ref


def process_bid_summary(bid_summary_df: pd.DataFrame, items_ws):
    items_df = bid_summary_df[['b', 'c', 'd', 'e', 'h']]
    categories = ["1 | Direct Labor", "2 | Indirect Labor", "3 | Labor Burdens", "4 | Contract Labor", "5 | Other Indirects", "6 | Material", "7 | Major Equipment", "8 | Subcontracts & Tech Services", "9 | Per Diem", "Z | Below the Line"]
    current_category = -1
    costCode_found = False
    curRow = 2
    costType = "DC | Direct job Cost"
    # keep track for going back later to update KPI data
    numDLRows = 0
    
    for row in items_df.itertuples(index=False):
        if not costCode_found and row[0] != "Cost Code":
            continue
        if row[0] == "Cost Code":
            costCode_found = True
            current_category += 1
            continue
        if row[0] == "Sub-Total":
            if current_category == 8:
                break
            costCode_found = False
            continue
        
        description = row[1]
        hours = 0.0
        cost = row[4]
        quantity = 0
        unit = ""
        
        match categories[current_category]:
            case "1 | Direct Labor" | "2 | Indirect Labor":
                hours = row[3]
            case "4 | Contract Labor":
                hours = row[2]
            case "6 | Material":
                if type(description) == float:
                    if row[2] == "Subtotal":
                        continue
                    else:
                        description = row[2]
            case "7 | Major Equipment":
                description = "Major Equipment"
            case "8 | Subcontracts & Tech Services":
                description = "Subcontracts"
            case "9 | Per Diem":
                description = "Per Diem"
                # Days * weeks * craft
                quantity = row[3] * row[2] * row[1]
                unit = "DAY"

        if hours == 0 and cost == 0:
            continue 
        if current_category == 0:
            numDLRows += 1
        items_ws.cell(row=curRow, column=1, value="GT | Grand total")
        items_ws.cell(row=curRow, column=2, value=costType)
        items_ws.cell(row=curRow, column=3, value=categories[current_category])
        items_ws.cell(row=curRow, column=4, value=description)
        items_ws.cell(row=curRow, column=5, value=quantity)
        items_ws.cell(row=curRow, column=6, value=unit)
        items_ws.cell(row=curRow, column=7, value=cost)
        items_ws.cell(row=curRow, column=8, value=hours)
        curRow += 1

    # Below the line stuff
    # get the index of the quote summary line
    table1_index = items_df.loc[items_df['b'] == 'QUOTE SUMMARY'].index[0] + 1
    while items_df.iloc[table1_index]['b'] != "QUOTE TOTAL (Capital Improvement)":
        row = items_df.iloc[table1_index]
        description = row["b"]
        if description == "OTHER INSERV TRADES":
            description = description + " " + row["c"]
        hours = 0.0
        cost = row["h"]
        quantity = 0
        unit = ""
        table1_index += 1
        if cost == 0:
            continue
        items_ws.cell(row=curRow, column=1, value="GT | Grand total")
        items_ws.cell(row=curRow, column=2, value=costType)
        items_ws.cell(row=curRow, column=3, value="Z | Below the Line")
        items_ws.cell(row=curRow, column=4, value=description)
        items_ws.cell(row=curRow, column=5, value=quantity)
        items_ws.cell(row=curRow, column=6, value=unit)
        items_ws.cell(row=curRow, column=7, value=cost)
        items_ws.cell(row=curRow, column=8, value=hours)
        curRow += 1
        
    # and the Other quote totals
    table2_index = items_df.loc[items_df['b'] == 'OTHER QUOTE TOTALS'].index[0] + 1
    while items_df.iloc[table2_index]['b'] != "QUOTE TOTAL (Capital Improvement) with Bid Bond and/or RMI Tax":
        row = items_df.iloc[table2_index]
        description = row["b"]
        hours = 0.0
        cost = row["h"]
        quantity = 0
        unit = ""
        table2_index += 1
        if cost == 0:
            continue
        items_ws.cell(row=curRow, column=1, value="GT | Grand total")
        items_ws.cell(row=curRow, column=2, value=costType)
        items_ws.cell(row=curRow, column=3, value="Z | Below the Line")
        items_ws.cell(row=curRow, column=4, value=description)
        items_ws.cell(row=curRow, column=5, value=quantity)
        items_ws.cell(row=curRow, column=6, value=unit)
        items_ws.cell(row=curRow, column=7, value=cost)
        items_ws.cell(row=curRow, column=8, value=hours)
        curRow += 1

    # Quantity and unit of the direct labor stuff
    # get the index of the Estimated key performance indicators line
    index = bid_summary_df.loc[bid_summary_df['b'] == 'ESTIMATED KEY PERFORMANCE INDICATORS (KPI)'].index[0] + 1
    for row in items_df.iloc[index:].itertuples(index=False):
        if row[1] == "Total":
            break
        description = row[1]
        quantity = row[2]
        unit = row[3]
        if quantity == 0:
            continue
        # 1 indexed
        idx = 1
        for desc in items_ws.iter_rows(min_row = 2, max_row = numDLRows, min_col = 4, max_col = 4, values_only=True):
            # starting row 2
            idx += 1
            if description == desc[0]:
                items_ws.cell(row= idx, column=5, value=quantity)
                items_ws.cell(row= idx, column=6, value=unit)
                break

def process_kpis(project_ws, bid_analysis_kpi_df):
    for row in bid_analysis_kpi_df.itertuples(index=False):
        key = row[1]
        if type(key) == str:
            if ref.mechanical_kpi_reference.get(key):
                project_ws.cell(row=ref.mechanical_kpi_reference[key], column=4, value=row[2])
        if key == "Total Quote $/Total General Contractor Cost":
            break

def process_bid_info(project_ws, bid_info_df: pd.DataFrame):
    # hardcode contact, brief Project Description
    # contact Info
    contact_df = bid_info_df.loc[13:18, ['b', 'c']]
    proj_desc = bid_info_df.loc[22, 'b']
    contact=""
    for row in contact_df.itertuples(index=False):
        contact += f"{row[1]}\n"
    project_ws.cell(row=22, column=4, value=contact)
    project_ws.cell(row=3, column=4, value=proj_desc)
    temp = bid_info_df.loc[6:11, ['g', 'j']]
    temp.columns = ['b', 'c']
    temp2 = bid_info_df.loc[13:19, ['g', 'j']]
    temp2.columns = ['b', 'c']
    temp3 = bid_info_df.loc[24:27, ['g', 'j']]
    temp3.columns = ['b', 'c']
    temp4 = bid_info_df.loc[30:34, ['g', 'j']]
    temp4.columns = ['b', 'c']
    
    reformatted_df = pd.concat([bid_info_df.loc[6:11, ['b', 'c']], bid_info_df.loc[30:36, ['b', 'c']], temp, temp2, temp3, temp4], axis=0)
    for row in reformatted_df.itertuples(index=False):
        key = row[0]
        if ref.bid_info_reference.get(key):
            project_ws.cell(row=ref.bid_info_reference[key], column=4, value=row[1])